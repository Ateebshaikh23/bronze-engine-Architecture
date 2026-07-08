"""
Excel Processor

Responsibilities
----------------
1. Read Excel directly from MinIO
2. Process workbook sheet by sheet
3. Batch rows before inserting
4. Preserve source data exactly as-is

No SQL
No MinIO connection
"""

from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from bronze_engine.common.logger import get_logger

logger = get_logger(__name__)


class ExcelProcessor:

    def __init__(self, postgres_loader):

        self.loader = postgres_loader

    # ==========================================================
    # Process Excel
    # ==========================================================

    def process(
        self,
        stream,
        dataset_name,
        folder_name,
        file_name,
        minio_path,
        batch_size=5000
    ):

        logger.info(f"Started processing : {file_name}")

        workbook = load_workbook(
            filename=BytesIO(stream.read()),
            read_only=True,
            data_only=True
        )

        total_rows = 0

        try:

            for sheet in workbook.worksheets:

                logger.info(f"Processing sheet : {sheet.title}")

                rows = sheet.iter_rows(values_only=True)

                try:
                    headers = next(rows)
                except StopIteration:
                    logger.warning(f"{sheet.title} is empty.")
                    continue

                headers = [
                    str(col).strip() if col is not None else f"column_{i+1}"
                    for i, col in enumerate(headers)
                ]

                batch = []

                for row in rows:

                    batch.append(row)

                    if len(batch) >= batch_size:

                        df = pd.DataFrame(
                            batch,
                            columns=headers
                        ).fillna("").astype(str)

                        self.loader.insert_dataframe(

                            dataframe=df,

                            table_name="bronze_excel",

                            dataset_name=dataset_name,

                            folder_name=folder_name,

                            file_name=file_name,

                            minio_path=minio_path

                        )

                        total_rows += len(batch)

                        logger.info(
                            f"{file_name} : {total_rows} rows loaded."
                        )

                        batch = []

                if batch:

                    df = pd.DataFrame(
                        batch,
                        columns=headers
                    ).fillna("").astype(str)

                    self.loader.insert_dataframe(

                        dataframe=df,

                        table_name="bronze_excel",

                        dataset_name=dataset_name,

                        folder_name=folder_name,

                        file_name=file_name,

                        minio_path=minio_path

                    )

                    total_rows += len(batch)

                    logger.info(
                        f"{file_name} : {total_rows} rows loaded."
                    )

            logger.info(f"{file_name} completed successfully.")

        finally:

            workbook.close()